import random
from openpyxl import Workbook

# Define word lists
subjects = ["The cat", "A person", "The robot", "My friend"]
verbs = ["eats", "runs", "jumps over", "thinks about"]
objects = ["a sandwich", "the wall", "a mystery", "the problem"]

# Generate a sentence
sentence = f"{random.choice(subjects)} {random.choice(verbs)} {random.choice(objects)}."

# Write to Excel
wb = Workbook()
ws = wb.active
ws["A1"] = "Generated Sentence"
ws["A2"] = sentence

# Save the file
wb.save("generated_sentences.xlsx")

print("Sentence saved to Excel!")
